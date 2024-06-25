import os
import openpyxl
import xlrd
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl_image_loader import SheetImageLoader
from xlrd.book import Book as XlrdBook

from .core import BaseParser, CustomParser
from .csv_file_parser import CsvFileParser
from openpyxl.cell.cell import Cell

from xlrd.sheet import Sheet


@CustomParser("excel", ["xls", "xlsx"])
class ExcelFileParser(CsvFileParser):
    """
    Subclass of FileParser specifically for parsing Excel files.
    Overrides load_file to load the workbook and extract_text to extract workbook's sheets.
    """

    def load_file(self):
        """
        Opens and reads the Excel file
        """
        _, file_extension = os.path.splitext(self.file_path)
        if file_extension == ".xlsx":
            self.file_content = openpyxl.load_workbook(filename=self.file_path, read_only=False, keep_vba=True)
        elif file_extension == ".xls":
            self.file_content = xlrd.open_workbook(self.file_path)
        else:
            raise ValueError(f"Unsupported file extension: {file_extension}. Only .xls and .xlsx are supported.")

    def extract_text(self) -> list[str]:
        """
        Extracts and returns text from an Excel file.
        """
        text = ""
        
        result_texts = []
        if isinstance(self.file_content, OpenpyxlWorkbook):
            for sheet_id, sheet_name in enumerate(self.file_content.sheetnames):
                current_sheet = self.file_content[sheet_name]
                image_loader = SheetImageLoader(current_sheet)
                # handle text cells
                processed_rows = []
                for row in current_sheet.iter_rows(values_only=False):
                    cells = []
                    for cell in row:
                        if image_loader.image_in(cell.coordinate):
                            title = f"image{sheet_id+1}-{cell.coordinate.lower()}"
                            cells.append(self.format_markdown_image(image_loader.get(cell.coordinate), title=title))
                        else:
                            cells.append(str(cell.value or ''))
                    processed_rows.append(cells)

                text = self.process_sheet(processed_rows, sheet_name)
                result_texts.append(text)

        elif isinstance(self.file_content, XlrdBook):
            for sheet_name in self.file_content.sheet_names():
                text = ""
                current_sheet = self.file_content[sheet_name]
                # handle text cells
                rows = current_sheet.get_rows()
                processed_rows = []
                for row in rows:
                    processed_rows.append([str(cell.value or '') for cell in row])
                    print(row)

                text = self.process_sheet(processed_rows, sheet_name)
                result_texts.append(text)
        else:
            raise ValueError("Unsupported workbook type.")

        return result_texts
    